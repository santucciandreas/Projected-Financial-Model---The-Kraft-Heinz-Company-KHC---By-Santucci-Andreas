"""
Kraft Heinz (KHC) — Interactive Financial Dashboard generator.

Reads the 'Insights' sheet of the projected financial model and renders a
self-contained, GitHub-Pages-ready index.html (Plotly via CDN).

Usage:
    python generate_dashboard.py model.xlsx index.html
"""

import json
import sys
from openpyxl import load_workbook

# Row map of the Insights sheet -> (key, english label, unit, format)
METRICS = {
    7:  ("net_sales",      "Net Sales",            "$mm", "money"),
    8:  ("sales_growth",   "Sales Growth",         "%",   "pct"),
    9:  ("ebitda",         "EBITDA",               "$mm", "money"),
    10: ("ebitda_margin",  "EBITDA Margin",        "%",   "pct"),
    11: ("net_income",     "Net Income",           "$mm", "money"),
    12: ("net_margin",     "Net Margin",           "%",   "pct"),
    16: ("fcf",            "Free Cash Flow",       "$mm", "money"),
    17: ("fcf_conversion", "FCF Conversion",       "%",   "pct"),
    21: ("roic",           "ROIC",                 "%",   "pct"),
    22: ("netdebt_ebitda", "Net Debt / EBITDA",    "x",   "mult"),
    26: ("eps",            "EPS",                  "$",   "dollar"),
    27: ("dividends",      "Dividends Paid",       "$mm", "money"),
    31: ("interest_cov",   "Interest Coverage",    "x",   "mult"),
    32: ("dividend_cov",   "Dividend Coverage",    "x",   "mult"),
    33: ("payout",         "Payout Ratio",         "%",   "pct"),
    34: ("debt_capital",   "Total Debt / Capital", "%",   "pct"),
}

YEAR_COLS = ["D", "E", "F", "G", "H", "I", "J", "K"]
YEARS = ["2022", "2023", "2024", "2025E", "2026E", "2027E", "2028E", "2029E"]
N_ACTUAL = 3  # 2022, 2023, 2024 are actuals; the rest projected


def read_data(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Insights"]
    data = {}
    for row, (key, label, unit, fmt) in METRICS.items():
        series = []
        for col in YEAR_COLS:
            v = ws[f"{col}{row}"].value
            series.append(v if isinstance(v, (int, float)) else None)
        data[key] = {"label": label, "unit": unit, "fmt": fmt, "values": series}
    return data


def build_html(data):
    payload = json.dumps({"years": YEARS, "n_actual": N_ACTUAL, "metrics": data})
    return HTML_TEMPLATE.replace("/*__DATA__*/", payload)


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Kraft Heinz (KHC) — Financial Model Insights</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Archivo:wght@500;600;700;800;900&family=Hanken+Grotesk:wght@400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js" charset="utf-8"></script>
<style>
  :root{
    --bg-0:#0f1830; --bg-1:#15223e; --navy:#1B2A4A;
    --panel:#1d2c4d; --panel-2:#223358;
    --line:rgba(255,255,255,.07); --line-2:rgba(255,255,255,.12);
    --gold:#C5A572; --gold-hi:#dcc196; --gold-dim:rgba(197,165,114,.16);
    --green:#46cf7c; --red:#e3725f; --periw:#7d97cf; --steel:#5d6f9c; --teal:#5cb6c7;
    --ink:#eef1f8; --body:#aeb7cb; --muted:#7a849b;
  }
  *{box-sizing:border-box;margin:0;padding:0}
  html{scroll-behavior:smooth}
  body{font-family:'Hanken Grotesk',system-ui,sans-serif;color:var(--body);line-height:1.5;
       background-color:#0f1830;
       background-image:radial-gradient(1200px 600px at 78% -8%, #1c2c4f 0%, transparent 60%), linear-gradient(180deg,#14213c 0%,#0f1830 70%);
       background-repeat:no-repeat;-webkit-font-smoothing:antialiased}
  .num{font-variant-numeric:tabular-nums;font-feature-settings:"tnum" 1}
  .wrap{max-width:1240px;margin:0 auto;padding:0 26px}
  em{font-style:normal;color:var(--gold)}

  /* ---------- TOP BAR ---------- */
  .topbar{border-bottom:1px solid var(--line);background:rgba(15,24,48,.55);backdrop-filter:blur(6px);
          position:sticky;top:0;z-index:30}
  .topbar .wrap{display:flex;align-items:center;justify-content:space-between;height:62px}
  .brand{display:flex;align-items:center;gap:13px}
  .mark{font-family:'Archivo',sans-serif;font-weight:800;font-size:15px;letter-spacing:.04em;color:#0f1830;
        background:var(--gold);padding:5px 9px;border-radius:6px;line-height:1}
  .brand-sub{font-size:13.5px;color:var(--body);letter-spacing:.01em}
  .brand-sub b{color:var(--ink);font-weight:700}
  .topbar-meta{display:flex;align-items:center;gap:9px;font-size:12px;color:var(--muted);letter-spacing:.04em}
  .pill{display:inline-flex;align-items:center;gap:6px;border:1px solid var(--line-2);border-radius:30px;padding:4px 11px;color:var(--body)}
  .pill::before{content:"";width:6px;height:6px;border-radius:50%;background:var(--green);box-shadow:0 0 0 3px rgba(70,207,124,.18)}

  /* ---------- HERO ---------- */
  .hero{padding:54px 0 8px}
  .eyebrow{font-family:'Archivo',sans-serif;font-size:11px;font-weight:700;letter-spacing:.22em;text-transform:uppercase;color:var(--gold)}
  .hero h1{font-family:'Archivo',sans-serif;font-weight:800;color:var(--ink);
           font-size:clamp(30px,4.6vw,52px);line-height:1.05;letter-spacing:-.015em;margin:16px 0 0;max-width:20ch}
  .hero .lede{margin-top:18px;max-width:74ch;font-size:clamp(14px,1.5vw,16.5px);color:var(--body)}

  /* ---------- KPI GRID ---------- */
  .kpis{padding:30px 0 8px;display:grid;grid-template-columns:repeat(4,1fr);gap:14px}
  .kpi{position:relative;background:linear-gradient(180deg,var(--panel-2),var(--panel));
       border:1px solid var(--line);border-radius:14px;padding:20px 22px 18px;overflow:hidden;
       transition:border-color .2s,transform .2s}
  .kpi::before{content:"";position:absolute;left:0;top:18px;bottom:18px;width:3px;border-radius:3px;background:var(--gold);opacity:.85}
  .kpi:hover{border-color:var(--line-2);transform:translateY(-2px)}
  .kpi .k-label{font-size:11.5px;font-weight:600;letter-spacing:.09em;text-transform:uppercase;color:var(--muted)}
  .kpi .k-val{font-family:'Archivo',sans-serif;font-weight:800;font-variant-numeric:tabular-nums;
              font-size:clamp(34px,3.6vw,46px);color:var(--ink);line-height:1;margin-top:12px;letter-spacing:-.02em}
  .kpi .k-val small{font-size:.5em;font-weight:700;color:var(--muted);margin-left:3px;letter-spacing:0}
  .kpi .k-foot{margin-top:13px;display:flex;align-items:center;gap:9px;font-size:12.5px;color:var(--muted)}
  .delta{display:inline-flex;align-items:center;gap:4px;font-weight:700;font-variant-numeric:tabular-nums;
         padding:2px 9px;border-radius:30px;font-size:12px}
  .delta.up{background:rgba(70,207,124,.14);color:var(--green)}
  .delta.down{background:rgba(227,114,95,.14);color:var(--red)}
  .delta .tri{font-size:9px;line-height:1}

  /* ---------- PANELS ---------- */
  main{padding:26px 0 10px}
  .grouphead{display:flex;align-items:center;gap:14px;margin:30px 2px 14px}
  .grouphead h2{font-family:'Archivo',sans-serif;font-weight:700;font-size:13px;letter-spacing:.16em;text-transform:uppercase;color:var(--body)}
  .grouphead .rule{flex:1;height:1px;background:var(--line)}

  .panel{background:linear-gradient(180deg,var(--panel),var(--navy));border:1px solid var(--line);
          border-radius:16px;overflow:hidden;margin-bottom:14px;transition:border-color .2s}
  .panel:hover{border-color:var(--line-2)}
  .panel-head{display:flex;align-items:center;gap:14px;padding:16px 22px;border-bottom:1px solid var(--line)}
  .panel-head .idx{font-family:'Archivo',sans-serif;font-weight:800;font-size:13px;color:var(--gold);letter-spacing:.04em}
  .panel-head h3{font-family:'Archivo',sans-serif;font-weight:700;font-size:16.5px;color:var(--ink);letter-spacing:-.01em}
  .panel-head .pull{margin-left:auto;font-size:12.5px;color:var(--body);font-variant-numeric:tabular-nums;
                    border:1px solid var(--line-2);border-radius:30px;padding:5px 13px;white-space:nowrap}
  .panel-head .pull b{color:var(--gold-hi);font-weight:700}
  .panel-grid{display:grid;grid-template-columns:288px 1fr}
  .panel-note{padding:24px 24px;border-right:1px solid var(--line)}
  .panel-note h4{font-family:'Archivo',sans-serif;font-weight:700;font-size:16px;color:var(--ink);margin-bottom:10px;line-height:1.25}
  .panel-note p{font-size:13.5px;color:var(--body)}
  .chart{min-height:330px;padding:14px 12px 6px}

  /* ---------- DATA TABLE ---------- */
  .data{margin-top:14px}
  details{background:linear-gradient(180deg,var(--panel),var(--navy));border:1px solid var(--line);border-radius:16px;overflow:hidden}
  summary{cursor:pointer;list-style:none;padding:18px 22px;font-family:'Archivo',sans-serif;font-weight:700;color:var(--ink);
          display:flex;align-items:center;justify-content:space-between;font-size:15px}
  summary::-webkit-details-marker{display:none}
  summary .hint{font-family:'Hanken Grotesk';font-weight:400;color:var(--muted);font-size:12.5px}
  summary .chev{transition:transform .25s ease;color:var(--gold)}
  details[open] summary .chev{transform:rotate(180deg)}
  .tbl-scroll{overflow-x:auto;border-top:1px solid var(--line)}
  table{border-collapse:collapse;width:100%;font-size:13px}
  th,td{padding:10px 15px;text-align:right;white-space:nowrap}
  th:first-child,td:first-child{text-align:left;position:sticky;left:0}
  thead th{background:#15233f;color:var(--body);font-family:'Archivo';font-weight:600;font-size:11.5px;letter-spacing:.05em}
  thead th.est{color:var(--gold)}
  thead th:first-child{background:#15233f}
  tbody td{font-variant-numeric:tabular-nums;color:var(--body);border-bottom:1px solid var(--line)}
  tbody td:first-child{color:var(--ink);font-weight:600;background:var(--navy)}
  tbody tr.grp td{background:rgba(255,255,255,.03);font-family:'Archivo';font-weight:700;color:var(--gold);
                  text-transform:uppercase;font-size:11px;letter-spacing:.07em}
  td.est{background:rgba(197,165,114,.05)}

  /* ---------- STATUS BAR ---------- */
  footer{margin-top:34px;border-top:1px solid var(--line);background:rgba(15,24,48,.5)}
  footer .wrap{display:flex;flex-wrap:wrap;gap:10px 22px;justify-content:space-between;align-items:center;
               padding:20px 26px;font-size:12.5px;color:var(--muted)}
  footer b{color:var(--body);font-weight:600}
  .legend-dot{display:inline-block;width:22px;height:0;border-top:2px dashed var(--gold);vertical-align:middle;margin-right:7px}

  /* ---------- RESPONSIVE ---------- */
  @media (max-width:900px){
    .kpis{grid-template-columns:repeat(2,1fr)}
    .panel-grid{grid-template-columns:1fr}
    .panel-note{border-right:none;border-bottom:1px solid var(--line)}
    .panel-head .pull{display:none}
  }
  @media (max-width:540px){
    .kpis{grid-template-columns:1fr}
    .wrap{padding-left:16px;padding-right:16px}
    .brand-sub{display:none}
  }
  @media (prefers-reduced-motion:reduce){*{scroll-behavior:auto!important;transition:none!important}}
</style>
</head>
<body>

<header class="topbar">
  <div class="wrap">
    <div class="brand"><span class="mark">KHC</span><span class="brand-sub"><b>Kraft Heinz</b> · Projected Financial Model</span></div>
    <div class="topbar-meta"><span style="font-size:12px;color:#7a849b">By <b style="color:#aeb7cb;font-weight:600">Andreas Santucci</b></span><span class="pill">Base case</span><span>FY2022 – 2029E</span></div>
  </div>
</header>

<section class="wrap hero">
  <span class="eyebrow">Model read · NASDAQ KHC</span>
  <h1 id="verdict"></h1>
  <p class="lede" id="lede"></p>
</section>

<section class="wrap kpis" id="kpis"></section>

<main class="wrap">
  <div class="grouphead"><h2>Performance &amp; balance sheet</h2><span class="rule"></span></div>
  <div id="sections"></div>

  <section class="data">
    <details>
      <summary>
        <span>Full data table <span class="hint">— 16 metrics · 2022–2029E</span></span>
        <span class="chev">▾</span>
      </summary>
      <div class="tbl-scroll"><table id="datatable"></table></div>
    </details>
  </section>
</main>

<footer>
  <div class="wrap">
    <span><span class="legend-dot"></span>Dashed gold line marks the actual / projected boundary in every chart.</span>
    <span><b>Kraft Heinz Co.</b> · Figures from the model's Insights sheet · Model &amp; dashboard by <b>Andreas Santucci</b> · May 2026</span>
  </div>
</footer>

<script>
const DATA = /*__DATA__*/;
const { years, n_actual, metrics } = DATA;
const M = metrics;

/* ---------- palette (tuned for the dark surface) ---------- */
const C = {gold:'#C5A572', goldHi:'#dcc196', periw:'#7d97cf', steel:'#6076a3',
           green:'#46cf7c', red:'#e3725f', teal:'#5cb6c7',
           grid:'rgba(255,255,255,0.07)', axis:'#8b94ab', label:'#aeb7cb'};
const FONT = "'Hanken Grotesk', sans-serif";
const DIVIDER_X = n_actual - 0.5;

const fmt = {
  money:v=>v==null?'–':(v<0?'-$':'$')+Math.abs(v).toLocaleString('en-US',{maximumFractionDigits:0}),
  moneyB:v=>v==null?'–':(v<0?'-$':'$')+(Math.abs(v)/1000).toFixed(1)+'B',
  pct:v=>v==null?'–':(v*100).toFixed(1)+'%',
  mult:v=>v==null?'–':v.toFixed(1)+'x',
  dollar:v=>v==null?'–':'$'+v.toFixed(2),
};
function fval(key,v){return fmt[M[key].fmt](v);}
function lastIdx(key){const a=M[key].values;for(let i=a.length-1;i>=0;i--)if(a[i]!=null)return i;return -1;}

/* ---------- hero copy ---------- */
document.getElementById('verdict').innerHTML = 'Flat sales, a <em>sharper engine</em>';
document.getElementById('lede').textContent =
  "Revenue holds near $26–27bn, so the model's real story is profitability and the balance sheet: "
  + "EBITDA margins rebuild from a 2024 trough toward ~22%, free cash flow turns durably positive, "
  + "and net leverage falls below 2.5× — reshaping Kraft Heinz into a stronger cash generator through 2029E.";

/* ---------- KPI cards ---------- */
const KPIS = [
  {key:'ebitda_margin', label:'EBITDA Margin', good:'up',   baseIdx:2, dtype:'pp',  cmp:'vs 2024'},
  {key:'fcf',           label:'Free Cash Flow', good:'up',   baseIdx:3, dtype:'pct', cmp:"'25E → '29E"},
  {key:'netdebt_ebitda',label:'Net Debt / EBITDA', good:'down', baseIdx:2, dtype:'x', cmp:'vs 2024'},
  {key:'eps',           label:'EPS', good:'up',              baseIdx:2, dtype:'pct', cmp:'vs 2024'},
];
const deltaFmt = {
  pp:(b,c)=>((c-b)*100>=0?'+':'')+((c-b)*100).toFixed(1)+'pp',
  x: (b,c)=>((c-b)>=0?'+':'')+(c-b).toFixed(1)+'x',
  pct:(b,c)=>((c-b)/Math.abs(b)*100>=0?'+':'')+((c-b)/Math.abs(b)*100).toFixed(0)+'%',
};
const kpiWrap = document.getElementById('kpis');
KPIS.forEach(k=>{
  const arr=M[k.key].values, li=lastIdx(k.key), cur=arr[li], base=arr[k.baseIdx];
  let dStr='—', cls='', rising=false;
  if(base!=null && cur!=null){ rising=cur>base; cls=((k.good==='up')?rising:!rising)?'up':'down'; dStr=deltaFmt[k.dtype](base,cur); }
  const tri = cls==='up'?(rising?'▲':'▼'):(cls==='down'?(rising?'▲':'▼'):'');
  const valStr=(k.key==='fcf')?fmt.moneyB(cur):fval(k.key,cur);
  kpiWrap.insertAdjacentHTML('beforeend', `
    <div class="kpi">
      <div class="k-label">${k.label}</div>
      <div class="k-val">${valStr}</div>
      <div class="k-foot"><span class="delta ${cls}"><span class="tri">${tri}</span>${dStr.replace(/^[+\-]/,'')}</span><span>${k.cmp}</span></div>
    </div>`);
});

/* ---------- sections ---------- */
const SECTIONS = [
  {n:'01', title:'Growth & Profitability', head:'Profitability, not growth, is the story',
   note:"Net sales stay essentially flat near $26–27bn. After a sharp 2024 dip, EBITDA margin rebuilds from ~6.5% toward roughly 22% by 2029E — the engine, not the top line, drives value.",
   pull:'EBITDA margin <b>6.5% → 22.2%</b>', render:chartGrowth},
  {n:'02', title:'Cash Generation', head:'A cash engine that finally turns over',
   note:"Free cash flow swings from negative in 2022–23 to consistently positive from 2024 onward, with conversion stabilizing near 47–49% — a far more reliable cash profile.",
   pull:'FCF conversion <b>~47–49%</b>', render:chartCash},
  {n:'03', title:'Efficiency & Solvency', head:'Deleveraging is the balance-sheet thesis',
   note:"ROIC climbs steadily toward ~5.4%, while net leverage collapses from a 2024 spike of 11× down to about 2.4× EBITDA — a clear, multi-year deleveraging path.",
   pull:'Net leverage <b>11.0× → 2.4×</b>', render:chartSolvency},
  {n:'04', title:'Shareholder Returns', head:'Steady dividend, recovering earnings',
   note:"Dividends hold steady near $1.9–2.0bn while EPS nearly triples off the 2024 trough, pulling the payout ratio back to a sustainable ~60%.",
   pull:'Payout ratio <b>149% → 60%</b>', render:chartReturns},
  {n:'05', title:'Risk Analysis', head:'Coverage strengthens across the board',
   note:"Interest coverage rises above 6× and dividend coverage holds above 1.3×, easing the balance-sheet risk that peaked in the weak 2024 results.",
   pull:'Interest coverage <b>0.8× → 6.0×</b>', render:chartRisk},
];

const layoutBase = () => ({
  font:{family:FONT, size:12, color:C.label},
  paper_bgcolor:'rgba(0,0,0,0)', plot_bgcolor:'rgba(0,0,0,0)',
  margin:{l:56,r:56,t:20,b:36}, hovermode:'x unified',
  hoverlabel:{font:{family:FONT,size:12,color:'#eef1f8'}, bgcolor:'#15233f', bordercolor:'rgba(255,255,255,.15)'},
  legend:{orientation:'h', y:-0.17, x:0.5, xanchor:'center', font:{size:11.5,color:C.label}},
  xaxis:{type:'category', tickfont:{size:11.5,color:C.label}, showgrid:false,
         linecolor:C.grid, ticks:'outside', tickcolor:C.grid},
  shapes:[
    {type:'rect', xref:'x', yref:'paper', x0:DIVIDER_X, x1:years.length-0.5, y0:0, y1:1,
     fillcolor:'rgba(197,165,114,0.07)', line:{width:0}, layer:'below'},
    {type:'line', xref:'x', yref:'paper', x0:DIVIDER_X, x1:DIVIDER_X, y0:0, y1:1,
     line:{color:C.gold, width:1.5, dash:'dash'}},
  ],
  annotations:[
    {x:DIVIDER_X, y:1.04, xref:'x', yref:'paper', text:'PROJECTED →', showarrow:false,
     font:{family:FONT,size:10,color:C.gold}, xanchor:'left'}
  ],
});
const yAxis = (title,extra={}) => Object.assign({
  title:{text:title, font:{size:11,color:C.axis}}, gridcolor:C.grid, zeroline:false,
  tickfont:{size:11,color:C.axis}, showline:false}, extra);
const CONFIG = {displayModeBar:false, responsive:true};

function chartGrowth(el){
  const t=[
    {type:'bar', name:'Net Sales', x:years, y:M.net_sales.values, marker:{color:C.periw},
     hovertemplate:'%{y:$,.0f}mm<extra>Net Sales</extra>'},
    {type:'bar', name:'EBITDA', x:years, y:M.ebitda.values, marker:{color:C.gold},
     hovertemplate:'%{y:$,.0f}mm<extra>EBITDA</extra>'},
    {type:'scatter', mode:'lines+markers', name:'EBITDA Margin', x:years, y:M.ebitda_margin.values,
     yaxis:'y2', line:{color:C.green,width:2.5}, marker:{size:6},
     hovertemplate:'%{y:.1%}<extra>EBITDA Margin</extra>'},
  ];
  const lay=Object.assign(layoutBase(),{barmode:'group',bargap:.32,bargroupgap:.08,
    yaxis:yAxis('$mm'),
    yaxis2:Object.assign(yAxis('Margin'),{overlaying:'y',side:'right',tickformat:'.0%',showgrid:false})});
  Plotly.newPlot(el,t,lay,CONFIG);
}
function chartCash(el){
  const colors=M.fcf.values.map(v=>v==null?C.axis:(v<0?C.red:C.green));
  const t=[
    {type:'bar', name:'Free Cash Flow', x:years, y:M.fcf.values, marker:{color:colors},
     hovertemplate:'%{y:$,.0f}mm<extra>FCF</extra>'},
    {type:'scatter', mode:'lines+markers', name:'FCF Conversion', x:years, y:M.fcf_conversion.values,
     yaxis:'y2', line:{color:C.gold,width:2.5}, marker:{size:6},
     hovertemplate:'%{y:.0%}<extra>Conversion</extra>'},
  ];
  const lay=Object.assign(layoutBase(),{yaxis:yAxis('$mm'),
    yaxis2:Object.assign(yAxis('Conversion'),{overlaying:'y',side:'right',tickformat:'.0%',showgrid:false})});
  Plotly.newPlot(el,t,lay,CONFIG);
}
function chartSolvency(el){
  const t=[
    {type:'scatter', mode:'lines+markers', name:'Net Debt / EBITDA', x:years, y:M.netdebt_ebitda.values,
     line:{color:C.periw,width:2.5}, marker:{size:6}, fill:'tozeroy', fillcolor:'rgba(125,151,207,0.10)',
     hovertemplate:'%{y:.1f}x<extra>Net Debt/EBITDA</extra>'},
    {type:'scatter', mode:'lines+markers', name:'ROIC', x:years, y:M.roic.values,
     yaxis:'y2', line:{color:C.gold,width:2.5}, marker:{size:6},
     hovertemplate:'%{y:.1%}<extra>ROIC</extra>'},
  ];
  const lay=Object.assign(layoutBase(),{yaxis:yAxis('Leverage (x)'),
    yaxis2:Object.assign(yAxis('ROIC'),{overlaying:'y',side:'right',tickformat:'.0%',showgrid:false})});
  Plotly.newPlot(el,t,lay,CONFIG);
}
function chartReturns(el){
  const t=[
    {type:'bar', name:'Dividends Paid', x:years, y:M.dividends.values, marker:{color:C.steel},
     hovertemplate:'%{y:$,.0f}mm<extra>Dividends</extra>'},
    {type:'scatter', mode:'lines+markers', name:'EPS', x:years, y:M.eps.values,
     yaxis:'y2', line:{color:C.gold,width:2.5}, marker:{size:6},
     hovertemplate:'$%{y:.2f}<extra>EPS</extra>'},
  ];
  const lay=Object.assign(layoutBase(),{yaxis:yAxis('$mm'),
    yaxis2:Object.assign(yAxis('EPS ($)'),{overlaying:'y',side:'right',tickprefix:'$',showgrid:false})});
  Plotly.newPlot(el,t,lay,CONFIG);
}
function chartRisk(el){
  const t=[
    {type:'scatter', mode:'lines+markers', name:'Interest Coverage', x:years, y:M.interest_cov.values,
     line:{color:C.periw,width:2.5}, marker:{size:6}, hovertemplate:'%{y:.1f}x<extra>Interest Cov.</extra>'},
    {type:'scatter', mode:'lines+markers', name:'Dividend Coverage', x:years, y:M.dividend_cov.values,
     line:{color:C.gold,width:2.5}, marker:{size:6}, hovertemplate:'%{y:.1f}x<extra>Dividend Cov.</extra>'},
    {type:'scatter', mode:'lines+markers', name:'Payout Ratio', x:years, y:M.payout.values,
     yaxis:'y2', line:{color:C.teal,width:2,dash:'dot'}, marker:{size:5},
     hovertemplate:'%{y:.0%}<extra>Payout</extra>'},
  ];
  const lay=Object.assign(layoutBase(),{yaxis:yAxis('Coverage (x)'),
    yaxis2:Object.assign(yAxis('Payout'),{overlaying:'y',side:'right',tickformat:'.0%',showgrid:false})});
  Plotly.newPlot(el,t,lay,CONFIG);
}

/* ---------- build sections ---------- */
const secWrap=document.getElementById('sections');
SECTIONS.forEach((s,i)=>{
  secWrap.insertAdjacentHTML('beforeend', `
    <section class="panel">
      <div class="panel-head"><span class="idx">${s.n}</span><h3>${s.title}</h3><span class="pull">${s.pull}</span></div>
      <div class="panel-grid">
        <div class="panel-note"><h4>${s.head}</h4><p>${s.note}</p></div>
        <div class="chart" id="chart-${i}"></div>
      </div>
    </section>`);
});

/* ---------- data table ---------- */
const GROUPS=[
  ['1 · Growth & Profitability',['net_sales','sales_growth','ebitda','ebitda_margin','net_income','net_margin']],
  ['2 · Cash Generation',['fcf','fcf_conversion']],
  ['3 · Efficiency & Solvency',['roic','netdebt_ebitda']],
  ['4 · Shareholder Returns',['eps','dividends']],
  ['5 · Risk Analysis',['interest_cov','dividend_cov','payout','debt_capital']],
];
const tbl=document.getElementById('datatable');
let head='<thead><tr><th>Metric</th><th>Unit</th>';
years.forEach((y,i)=>head+=`<th class="${i>=n_actual?'est':''}">${y}</th>`);
head+='</tr></thead>';
let body='<tbody>';
GROUPS.forEach(([g,keys])=>{
  body+=`<tr class="grp"><td colspan="${years.length+2}">${g}</td></tr>`;
  keys.forEach(k=>{
    body+=`<tr><td>${M[k].label}</td><td style="color:var(--muted)">${M[k].unit}</td>`;
    M[k].values.forEach((v,i)=>body+=`<td class="${i>=n_actual?'est':''}">${fval(k,v)}</td>`);
    body+='</tr>';
  });
});
body+='</tbody>';
tbl.innerHTML=head+body;

/* ---------- lazy render ---------- */
const io=new IntersectionObserver((entries,obs)=>{
  entries.forEach(e=>{ if(e.isIntersecting){ const i=+e.target.dataset.idx;
    SECTIONS[i].render(document.getElementById('chart-'+i)); obs.unobserve(e.target); } });
},{rootMargin:'140px'});
SECTIONS.forEach((s,i)=>{const c=document.getElementById('chart-'+i);c.dataset.idx=i;io.observe(c);});
window.addEventListener('resize',()=>{document.querySelectorAll('.chart').forEach(c=>{if(c.data)Plotly.Plots.resize(c);});});
</script>
</body>
</html>
"""


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "model.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "index.html"
    data = read_data(src)
    html = build_html(data)
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {out} ({len(html):,} bytes) from {src}")


if __name__ == "__main__":
    main()
